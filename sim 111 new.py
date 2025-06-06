import os
from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import messagebox, ttk
from datetime import datetime

BARANG_FILE = 'TokoBangunan.xlsx'
JURNAL_FILE = 'jurnal.xlsx'

def inisialisasi_file():
    if not os.path.exists(BARANG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Kode Barang', 'Nama Barang', 'Harga Jual', 'Harga Beli', 'Stok'])
        wb.save(BARANG_FILE)
    if not os.path.exists(JURNAL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Tanggal', 'Keterangan', 'Akun', 'Debit', 'Kredit'])
        wb.save(JURNAL_FILE)

def input_barang(kode, nama, harga_jual, harga_beli, stok):
    wb = load_workbook(BARANG_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0].lower() == kode.lower():
            messagebox.showerror("Error", "Kode barang sudah ada.")
            return
        if row[1] and row[1].lower() == nama.lower():
            messagebox.showerror("Error", "Nama barang sudah ada.")
            return
    ws.append([kode, nama, harga_jual, harga_beli, stok])
    wb.save(BARANG_FILE)
    
    # Catat pembelian persediaan awal ke jurnal
    total_persediaan = harga_beli * stok
    jurnal_entry(
        keterangan=f"Pembelian persediaan {nama} ({kode})",
        akun_debit="Persediaan Barang",
        debit=total_persediaan,
        akun_kredit="Kas",
        kredit=total_persediaan
    )
    
    messagebox.showinfo("Sukses", "Barang berhasil ditambahkan dan dicatat di jurnal.")

def tambah_stock_Persediaan(kode, tambahan):
    wb = load_workbook(BARANG_FILE)
    ws = wb.active
    ditemukan = False
    harga_beli = 0
    nama_barang = ""
    
    for row in ws.iter_rows(min_row=2):
        if row[0].value and row[0].value.lower() == kode.lower():
            row[4].value += tambahan
            harga_beli = row[3].value
            nama_barang = row[1].value
            ditemukan = True
            break
    
    if ditemukan:
        wb.save(BARANG_FILE)
        
        # Catat penambahan stok ke jurnal
        total_pembelian = harga_beli * tambahan
        jurnal_entry(
            keterangan=f"Penambahan stok {nama_barang} ({kode}) - {tambahan} unit",
            akun_debit="Persediaan Barang",
            debit=total_pembelian,
            akun_kredit="Kas",
            kredit=total_pembelian
        )
        
        messagebox.showinfo("Sukses", f"Stock berhasil ditambahkan dan dicatat di jurnal. Total: Rp{total_pembelian:,.2f}")
    else:
        messagebox.showerror("Error", "Kode barang tidak ditemukan.")

def jurnal_entry(keterangan, akun_debit=None, debit=0, akun_kredit=None, kredit=0):
    """Fungsi untuk mencatat transaksi ke jurnal umum"""
    wb = load_workbook(JURNAL_FILE)
    ws = wb.active
    tanggal = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Jika ada akun debit
    if akun_debit and debit > 0:
        ws.append([tanggal, keterangan, akun_debit, debit, 0])
    
    # Jika ada akun kredit
    if akun_kredit and kredit > 0:
        ws.append([tanggal, keterangan, akun_kredit, 0, kredit])
    
    wb.save(JURNAL_FILE)

def input_beban(keterangan, amount):
    """Fungsi untuk mencatat beban ke jurnal. Beban dicatat sebagai debit ke akun Beban, kredit ke Kas."""
    if amount <= 0:
        messagebox.showerror("Error", "Jumlah beban harus lebih dari 0.")
        return
    jurnal_entry(
        keterangan=keterangan,
        akun_debit="Beban",
        debit=amount,
        akun_kredit="Kas",
        kredit=amount
    )
    messagebox.showinfo("Sukses", "Beban berhasil dicatat di jurnal.")

def Persediaan_barang():
    wb = load_workbook(BARANG_FILE)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Pastikan ada data
            data.append((row[0], row[1], row[2], row[3], row[4]))
    return data

# Fungsi baru untuk persediaan barang khusus pelanggan (tanpa harga beli)
def Persediaan_barang_pelanggan():
    wb = load_workbook(BARANG_FILE)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Pastikan ada data
            # Hanya ambil kode, nama, harga jual, dan stok (tanpa harga beli)
            data.append((row[0], row[1], row[2], row[4]))
    return data

def beli_barang(kode_barang, jumlah):
    if jumlah <= 0:
        messagebox.showerror("Error", "Jumlah harus lebih dari 0.")
        return False
    
    wb_barang = load_workbook(BARANG_FILE)
    ws_barang = wb_barang.active
    ditemukan = False
    
    for row in ws_barang.iter_rows(min_row=2):
        if row[0].value and row[0].value.lower() == kode_barang.lower():
            stok = row[4].value
            harga_jual = row[2].value
            harga_beli = row[3].value
            nama_barang = row[1].value
            
            if jumlah > stok:
                messagebox.showerror("Error", f"Stok tidak cukup. Stok tersedia: {stok}")
                return False
            
            # Kurangi stok
            row[4].value -= jumlah
            ditemukan = True
            break
    
    if not ditemukan:
        messagebox.showerror("Error", "Kode barang tidak ditemukan.")
        return False
    
    wb_barang.save(BARANG_FILE)
    
    # Hitung total penjualan dan HPP
    total_penjualan = harga_jual * jumlah
    harga_pokok_penjualan = harga_beli * jumlah  # Gunakan harga beli untuk HPP
    
    # Catat ke jurnal umum (sistem double entry)
    keterangan = f"Penjualan {kode_barang} ({kode_barang}) - {jumlah} unit"
    
    # 1. Kas bertambah (Debit)
    jurnal_entry(
        keterangan=keterangan,
        akun_debit="Kas",
        debit=total_penjualan
    )
    
    # 2. Pendapatan bertambah (Kredit)
    jurnal_entry(
        keterangan=keterangan,
        akun_kredit="Pendapatan Penjualan",
        kredit=total_penjualan
    )
    
    # 3. Harga Pokok Penjualan (Debit)
    jurnal_entry(
        keterangan=f"HPP - {keterangan}",
        akun_debit="Harga Pokok Penjualan",
        debit=harga_pokok_penjualan
    )
    
    # 4. Persediaan berkurang (Kredit)
    jurnal_entry(
        keterangan=f"HPP - {keterangan}",
        akun_kredit="Persediaan Barang",
        kredit=harga_pokok_penjualan
    )
    
    messagebox.showinfo("Sukses", f"Pembelian berhasil! Total bayar: Rp{total_penjualan:,.2f}")
    return True

def Jurnal_Umum():
    """Mengambil data jurnal umum dan menghitung total saldo"""
    wb = load_workbook(JURNAL_FILE)
    ws = wb.active
    data = []
    total_debit = 0
    total_kredit = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Pastikan ada data
            debit = row[3] if row[3] else 0
            kredit = row[4] if row[4] else 0
            data.append((row[0], row[1], row[2], debit, kredit))
            total_debit += debit
            total_kredit += kredit
    
    return data, total_debit, total_kredit


def laporan_laba_rugi_data():
    """
    Menghitung laporan laba rugi dengan klasifikasi akun yang benar sesuai standar akuntansi sederhana.
    Pendapatan dicatat sebagai kredit, beban sebagai debit. HPP termasuk beban pokok penjualan.
    """
    wb = load_workbook(JURNAL_FILE)
    ws = wb.active
    
    total_pendapatan = 0
    total_hpp = 0
    total_beban = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            akun = row[2]
            debit = row[3] if row[3] else 0
            kredit = row[4] if row[4] else 0
            
            if akun == "Pendapatan Penjualan":
                total_pendapatan += kredit
            elif akun == "Harga Pokok Penjualan":
                total_hpp += debit
            elif akun == "Beban":
                total_beban += debit
    
    laba_rugi = total_pendapatan - total_hpp - total_beban
    
    return total_pendapatan, total_hpp, total_beban, laba_rugi

class Aplikasi:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistem Manajemen Toko Bangunan Putra")
        self.root.geometry("1000x700")
        self.root.minsize(1000, 700)
        self.login_frame()
        self.root.configure(bg="#ECEAD9")

    def clear_root(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def login_frame(self):
        self.clear_root()
        frame = Frame(self.root, bg="#8E9FBC")
        frame.pack(pady=100)
        Label(frame, text="Username:", font=("Arial", 12), bg="#ECEAD9").grid(row=0, column=0, pady=5, padx=5)
        self.username_entry = Entry(frame, font=("Arial", 12))
        self.username_entry.grid(row=0, column=1)
        Label(frame, text="Password:", font=("Arial", 12), bg="#ECEAD9").grid(row=1, column=0, pady=5, padx=5)
        self.password_entry = Entry(frame, show='*', font=("Arial", 12))
        self.password_entry.grid(row=1, column=1)
        Button(frame, text="Login", command=self.cek_login, font=("Arial", 12), bg="#8E9FBC", fg="white").grid(row=2, columnspan=2, pady=10)

    def cek_login(self):
        user = self.username_entry.get()
        pw = self.password_entry.get()
        if user == 'admin' and pw == 'bangunan1':
            self.admin_menu()
        elif user == 'pelanggan' and pw == 'pelanggan1':
            self.pelanggan_menu()
        else:
            messagebox.showerror("Login Gagal", "Username atau Password salah")

    def admin_menu(self):
        self.clear_root()
        main_frame = Frame(self.root, bg="#8E9FBC")
        main_frame.pack(pady=20, padx=20, fill=BOTH, expand=True)

        desc_frame = Frame(main_frame, bg="#ECEAD9", relief="ridge", bd=2)
        desc_frame.pack(pady=(0, 20), fill=X)

        Label(desc_frame, text="Selamat datang di Sistem Manajemen Toko Bangunan Putra!",
              font=("Arial", 16, "bold"), bg="#ECEAD9", fg="#2C3E50").pack(pady=10)

        desc_text = ("Aplikasi ini dirancang untuk membantu Anda dalam mengelola operasi toko bangunan dengan efisien.\n"
                     "Sistem ini menyediakan fitur-fitur lengkap untuk membantu mencatat persediaan, penjualan, dan jurnal akuntansi.")
        Label(desc_frame, text=desc_text, font=("Arial", 11), bg="#ECEAD9", fg="#34495E",
              wraplength=700, justify=CENTER).pack(pady=(0, 10), padx=20)

        button_frame = Frame(main_frame, bg="#ECEAD9")
        button_frame.pack(fill=BOTH, expand=True)

        def buat_tombol(teks, perintah, warna="#FFC9D6"):
            Button(button_frame, text=teks, command=perintah,
                   font=("Arial", 12), height=2, bg=warna, fg="black").pack(pady=5, fill=X)

        buat_tombol("Input Barang", self.form_tambah_barang)
        buat_tombol("Tambah Stock Persediaan", self.form_tambah_stok)
        buat_tombol("Persediaan", self.tampilkan_barang)
        buat_tombol("Jurnal Umum", self.tampilkan_jurnal)
        buat_tombol("Input Beban", self.form_input_beban)
        buat_tombol("Laporan Laba Rugi", self.tampilkan_laporan_laba_rugi)
        buat_tombol("Logout", self.login_frame, warna="#F6EFE5")

    def form_input_beban(self):
        win = Toplevel(self.root)
        win.title("Input Beban")
        win.geometry("350x200")

        Label(win, text="Keterangan Beban").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        keterangan_entry = Entry(win, width=30)
        keterangan_entry.grid(row=0, column=1, padx=10, pady=5)

        Label(win, text="Jumlah Beban (Rp)").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        jumlah_entry = Entry(win, width=30)
        jumlah_entry.grid(row=1, column=1, padx=10, pady=5)

        def simpan_beban():
            try:
                keterangan = keterangan_entry.get().strip()
                jumlah = float(jumlah_entry.get())
                if not keterangan:
                    messagebox.showerror("Error", "Keterangan tidak boleh kosong.")
                    return
                if jumlah <= 0:
                    messagebox.showerror("Error", "Jumlah beban harus lebih dari 0.")
                    return
                input_beban(keterangan, jumlah)
                win.destroy()
            except ValueError:
                messagebox.showerror("Error", "Jumlah harus angka.")

        Button(win, text="Simpan Beban", command=simpan_beban, bg="#8E9FBC", fg="white").grid(row=2, columnspan=2, pady=10)

    def tampilkan_laporan_laba_rugi(self):
        win = Toplevel(self.root)
        win.title("Laporan Laba Rugi Sederhana")
        win.geometry("500x350")

        total_pendapatan, total_hpp, total_beban, laba_rugi = laporan_laba_rugi_data()

        # Use TreeView for aligned debit / kredit columns
        columns = ("Akun", "Debit", "Kredit")
        tree = ttk.Treeview(win, columns=columns, show='headings', height=7)
        tree.heading("Akun", text="Akun")
        tree.heading("Debit", text="Debit")
        tree.heading("Kredit", text="Kredit")

        tree.column("Akun", anchor=W, width=250)
        tree.column("Debit", anchor=E, width=100)
        tree.column("Kredit", anchor=E, width=100)

        # Insert rows - Pendapatan Penjualan as Kredit (right), HPP and Beban as Debit (left)
        tree.insert("", END, values=("Pendapatan Penjualan", "", f"Rp{total_pendapatan:,.2f}"))
        tree.insert("", END, values=("Harga Pokok Penjualan", f"Rp{total_hpp:,.2f}", ""))
        tree.insert("", END, values=("Beban", f"Rp{total_beban:,.2f}", ""))
        tree.insert("", END, values=("", "", ""))  # Empty row for spacing
        # Net profit or loss row with styling
        if laba_rugi >= 0:
            tree.insert("", END, values=("Laba Bersih", "", f"Rp{laba_rugi:,.2f}"))
        else:
            # Show loss as debit to left side with minus sign
            tree.insert("", END, values=("Rugi Bersih", f"Rp{-laba_rugi:,.2f}", ""))

        tree.pack(pady=20, fill=X, padx=20)

    def form_tambah_barang(self):
        win = Toplevel(self.root)
        win.title("Input Barang")
        win.geometry("400x350")
        labels = ["Kode Barang", "Nama Barang", "Harga Jual", "Harga Beli", "Jumlah Unit"]
        entries = []
        for i, text in enumerate(labels):
            Label(win, text=text).grid(row=i, column=0, padx=10, pady=5, sticky="w")
            ent = Entry(win, width=25)
            ent.grid(row=i, column=1, padx=10, pady=5)
            entries.append(ent)
        
        def simpan():
            try:
                kode = entries[0].get().strip()
                nama = entries[1].get().strip()
                harga_jual = float(entries[2].get())
                harga_beli = float(entries[3].get())
                stok = int(entries[4].get())
                
                if not kode or not nama:
                    messagebox.showerror("Error", "Kode dan nama barang tidak boleh kosong")
                    return
                if harga_jual < 0 or harga_beli < 0 or stok < 0:
                    messagebox.showerror("Error", "Harga dan jumlah unit harus positif")
                    return
                
                input_barang(kode, nama, harga_jual, harga_beli, stok)
                win.destroy()
            except ValueError:
                messagebox.showerror("Error", "Harga harus angka dan jumlah unit harus integer")
        
        Button(win, text="Simpan", command=simpan, bg="#8E9FBC", fg="white").grid(row=5, columnspan=2, pady=10)

    def form_tambah_stok(self):
        win = Toplevel(self.root)
        win.title("Tambah Stock Persediaan")
        win.geometry("300x200")
        Label(win, text="Kode Barang").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        kode = Entry(win, width=25)
        kode.grid(row=0, column=1, padx=10, pady=5)
        Label(win, text="Jumlah Unit").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        tambahan = Entry(win, width=25)
        tambahan.grid(row=1, column=1, padx=10, pady=5)
        
        def simpan():
            try:
                kode_barang = kode.get().strip()
                jumlah = int(tambahan.get())
                
                if not kode_barang:
                    messagebox.showerror("Error", "Kode barang tidak boleh kosong")
                    return
                if jumlah <= 0:
                    messagebox.showerror("Error", "Jumlah Unit harus > 0")
                    return
                
                tambah_stock_Persediaan(kode_barang, jumlah)
                win.destroy()
            except ValueError:
                messagebox.showerror("Error", "Jumlah Unit harus integer")
        
        Button(win, text="Tambah", command=simpan, bg="#8E9FBC", fg="white").grid(row=2, columnspan=2, pady=10)

    def tampilkan_barang(self):
        win = Toplevel(self.root)
        win.title("Daftar Persediaan")
        win.geometry("1000x700")
        
        data = Persediaan_barang()
        
        tree = ttk.Treeview(win, columns=("Kode", "Nama", "Harga Jual", "Harga Beli", "Stok"), show='headings')
        tree.heading("Kode", text="Kode Barang")
        tree.heading("Nama", text="Nama Barang")
        tree.heading("Harga Jual", text="Harga Jual")
        tree.heading("Harga Beli", text="Harga Beli")
        tree.heading("Stok", text="Stok")
        
        # Atur lebar kolom
        tree.column("Kode", width=120)
        tree.column("Nama", width=280)
        tree.column("Harga Jual", width=150)
        tree.column("Harga Beli", width=150)
        tree.column("Stok", width=100)
        
        total_nilai_persediaan = 0
        for row in data:
            nilai_persediaan = row[3] * row[4]  # Harga beli * stok
            total_nilai_persediaan += nilai_persediaan
            tree.insert('', END, values=(row[0], row[1], f"Rp{row[2]:,.2f}", f"Rp{row[3]:,.2f}", row[4]))
        
        tree.pack(fill=BOTH, expand=True, pady=10)
        
        # Tampilkan total nilai persediaan
        frame_total = Frame(win)
        frame_total.pack(pady=10)
        Label(frame_total, text=f"Total Nilai Persediaan: Rp{total_nilai_persediaan:,.2f}", 
              font=('Arial', 12, 'bold')).pack()

    # Fungsi baru untuk menampilkan persediaan bagi pelanggan (tanpa harga beli)
    def tampilkan_barang_pelanggan(self):
        win = Toplevel(self.root)
        win.title("Daftar Barang")
        win.geometry("700x600")
        
        # Menggunakan fungsi khusus untuk pelanggan
        data = Persediaan_barang_pelanggan()
        
        tree = ttk.Treeview(win, columns=("Nama", "Harga Jual", "Stok"), show='headings')
        tree.heading("Nama", text="Nama Barang")
        tree.heading("Harga Jual", text="Harga")
        tree.heading("Stok", text="Stok Tersedia")
        
        # Atur lebar kolom
        tree.column("Nama", width=280)
        tree.column("Harga Jual", width=150)
        tree.column("Stok", width=100)
        
        for row in data:
            tree.insert('', END, values=(row[0], row[1], f"Rp{row[2]:,.2f}", row[3]))
        
        tree.pack(fill=BOTH, expand=True, pady=10)
        
        # Tambahkan label informasi
        Label(win, text="Silahkan catat kode barang untuk melakukan pembelian", 
              font=('Arial', 11), fg="#34495E").pack(pady=10)

    def tampilkan_jurnal(self):
        win = Toplevel(self.root)
        win.title("Jurnal Umum")
        win.geometry("1200x700")
        
        data, total_debit, total_kredit = Jurnal_Umum()
        
        tree = ttk.Treeview(win, columns=("Tanggal", "Keterangan", "Akun", "Debit", "Kredit"), show='headings')
        tree.heading("Tanggal", text="Tanggal")
        tree.heading("Keterangan", text="Keterangan")
        tree.heading("Akun", text="Akun")
        tree.heading("Debit", text="Debit")
        tree.heading("Kredit", text="Kredit")
        
        # Atur lebar kolom dan alignments
        tree.column("Tanggal", width=150)
        tree.column("Keterangan", width=350)
        tree.column("Akun", width=200)
        tree.column("Debit", width=150, anchor=E)
        tree.column("Kredit", width=150, anchor=E)
        
        for row in data:
            debit_str = f"Rp{row[3]:,.2f}" if row[3] > 0 else ""
            kredit_str = f"Rp{row[4]:,.2f}" if row[4] > 0 else ""
            # Indent akun name if it's a kredit (credit)
            akun_display = row[2]
            if row[4] > 0:  # kredit entry
                akun_display = "     " + akun_display  # Indent by 5 spaces
                
            tree.insert('', END, values=(row[0], row[1], akun_display, debit_str, kredit_str))
        
        tree.pack(fill=BOTH, expand=True, pady=10)
        
        # Frame untuk total
        frame_total = Frame(win)
        frame_total.pack(pady=10)
        
        Label(frame_total, text=f"Total Debit: Rp{total_debit:,.2f}", 
              font=('Arial', 12, 'bold')).pack(side=LEFT, padx=20)
        Label(frame_total, text=f"Total Kredit: Rp{total_kredit:,.2f}", 
              font=('Arial', 12, 'bold')).pack(side=LEFT, padx=20)
        
        # Validasi keseimbangan
        if abs(total_debit - total_kredit) < 0.01:  # Toleransi untuk floating point
            Label(frame_total, text="✓ SEIMBANG", fg="green", 
                  font=('Arial', 12, 'bold')).pack(side=LEFT, padx=20)
        else:
            Label(frame_total, text="✗ TIDAK SEIMBANG", fg="red", 
                  font=('Arial', 12, 'bold')).pack(side=LEFT, padx=20)


    def pelanggan_menu(self):
        self.clear_root()
        frame = Frame(self.root, bg="#8E9FBC")
        frame.pack(pady=20, fill=BOTH, expand=True)
        
        desc_frame = Frame(frame, bg="#ECEAD9", relief="ridge", bd=2)
        desc_frame.pack(pady=(0, 20), fill=X, padx=20)
        
        Label(desc_frame, text="Selamat Datang Pelanggan Toko Bangunan Putra!",
              font=("Arial", 16, "bold"), bg="#ECEAD9", fg="#2C3E50").pack(pady=10)
        
        def buat_btn(teks, cmd, warna="#766C28"):
            Button(frame, text=teks, command=cmd, font=("Arial", 12), 
                   height=2, bg=warna, fg="white").pack(pady=5, fill=X, padx=20)
        
        # Ubah untuk menggunakan fungsi tampilkan_barang_pelanggan
        buat_btn("Lihat Persediaan", self.tampilkan_barang_pelanggan, warna="#FFC9D6")
        buat_btn("Beli Barang", self.form_beli_barang, warna="#FFC9D6")
        buat_btn("Logout", self.login_frame, warna="#F6EFE5")

    def form_beli_barang(self):
        win = Toplevel(self.root)
        win.title("Beli Barang")
        win.geometry("300x200")
        
        Label(win, text="Kode Barang").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        kode = Entry(win, width=25)
        kode.grid(row=0, column=1, padx=10, pady=5)
        
        Label(win, text="Jumlah").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        jumlah = Entry(win, width=25)
        jumlah.grid(row=1, column=1, padx=10, pady=5)
        
        def beli():
            try:
                kode_barang = kode.get().strip()
                jml = int(jumlah.get())
                
                if not kode_barang:
                    messagebox.showerror("Error", "Kode barang tidak boleh kosong")
                    return
                if jml <= 0:
                    messagebox.showerror("Error", "Jumlah harus lebih dari 0")
                    return
                
                if beli_barang(kode_barang, jml):
                    win.destroy()
            except ValueError:
                messagebox.showerror("Error", "Jumlah harus integer")
        
        Button(win, text="Beli", command=beli, bg="#766C28", fg="white").grid(row=2, columnspan=2, pady=10)

if __name__ == "__main__":
    inisialisasi_file()
    root = Tk()
    app = Aplikasi(root)
    root.mainloop()


